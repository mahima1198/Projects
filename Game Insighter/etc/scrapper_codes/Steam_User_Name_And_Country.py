import requests
from datetime import datetime
import pycountry
import json
import openpyxl
import time
user_profile_url="http://api.steampowered.com/ISteamUser/GetPlayerSummaries/v0002/"
key='4F1B5E492903F0967512AD3AE03BA971'
wb1=openpyxl.load_workbook("games_id_names.xlsx")
sh1=wb1.active
for i in range(11,14): 
    game_name=(sh1.cell(row=i,column=1)).value
    filename=game_name+".xlsx"
    print("Game :: ",game_name)
    wb2=openpyxl.load_workbook(filename)
    sh2=wb2.active
    for j in range(2,sh2.max_row+1):
        #print(j)
        user_steam_id=int((sh2.cell(row=j,column=2)).value)
        user_profile_param={'key':key,'steamids':user_steam_id}
        r2=requests.get(url=user_profile_url,params=user_profile_param)
        user_data=r2.json()
        user_name=user_data['response']['players'][0].get('personaname')
        user_country_code=user_data['response']['players'][0].get('loccountrycode')
        sh2.cell(row=j,column=3).value=user_name
        print(user_name)
        if(user_country_code!=None):
            x=pycountry.countries.get(alpha_2=user_country_code)
            if(x==None):
                sh2.cell(row=j,column=4).value="null"
            else:
                sh2.cell(row=j,column=4).value=x.name
        else:
            sh2.cell(row=j,column=4).value="null"
        wb2.save(filename)
        if((j % 30) == 0 ):
            time.sleep(7)
    print("names and country added in : ",filename)
print("Finished")
