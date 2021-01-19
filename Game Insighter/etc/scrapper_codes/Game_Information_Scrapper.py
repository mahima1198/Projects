import requests
from datetime import datetime
import pycountry
import json
import openpyxl
import time
from bs4 import BeautifulSoup
row=2
def retrive_data(filename,game_link):
    if(filename=="Far Cry 5 data.xlsx" or filename=="Tom Clancy's Rainbow Six Siege data.xlsx"):
        return None
    global row
    
    #Using Scrapper
    r=requests.get(game_link)
    c=r.content
    soup=BeautifulSoup(c,"html.parser")
    
    #Extracting Game Title
    game_title=soup.find_all("div",{"class":"apphub_AppName"})[0]
    sh2.cell(row=row,column=1).value=game_title.text

    #Game Release Date
    game_rdate=soup.find_all("div",{"class":"date"})[0]
    sh2.cell(row=row,column=2).value=game_rdate.text

    #Game Publisher and Developer
    game_publisher=soup.find_all("div",{"class":"summary column"})[2]
    game_publisher=(game_publisher.text).replace("\n","")
    sh2.cell(row=row,column=3).value=game_publisher

    # Game Price
    game_price=soup.find_all("div",{"class":"game_purchase_price price"})[0]
    game_price=(game_price.text).replace("\t","")
    game_price=game_price.replace("\r","")
    game_price=game_price.replace("\n","")
    sh2.cell(row=row,column=4).value=game_price

    #Game Language
    table = soup.find('table',{'class':'game_language_options'})
    game_languages=""
    table_cols=table.find_all('td',{'class':'ellipsis'})
    for i in table_cols:
        z=i.text
        z=z.replace("\t","")
        z=z.replace("\n","")
        z=z.replace("\r","")
        game_languages=game_languages + (z+", ")
    sh2.cell(row=row,column=5).value=game_languages
    
    #Game Description
    game_description=soup.find('div',{'id':'game_area_description'})
    game_description=game_description.text
    game_description=game_description.replace("About This Game","")
    game_description=game_description.replace("\t","")
    game_description=game_description.replace("\r","")
    game_description=game_description.replace("\n","")
    sh2.cell(row=row,column=6).value=game_description
    
    #Mature Content Description
    game_mcd=soup.find('div',{'id':'game_area_content_descriptors'})
    if(game_mcd!=None):
        game_mcd=game_mcd.text
        game_mcd=game_mcd.replace("Mature Content Description","")
        game_mcd=game_mcd.replace("\r","")
        game_mcd=game_mcd.replace("\t","")
        game_mcd=game_mcd.replace("\n","")
        sh2.cell(row=row,column=7).value=game_mcd

    #System Requirements
    game_data=soup.find('ul',{'class':'bb_ul'})
    game_data=game_data.find_all('li')
    game_sysreq=""
    for x in game_data:
        temp=x.text
        game_sysreq=game_sysreq+ (temp+" $ ")
    sh2.cell(row=row,column=8).value=game_sysreq

    #Meta Critic Score
    game_mscore=soup.find('div',{'class':'score high'})
    #print(type(game_mscore))
    if(game_mscore!=None):
        
        game_score=game_mscore.text
        game_score=game_score.replace("\r","")
        game_score=game_score.replace("\n","")
        game_score=game_score.replace("\t","")
        sh2.cell(row=row,column=9).value=game_score

        #Meta Critic Score Link
        game_mlink=soup.find('div',{'id':'game_area_metalink'})
        game_mlink=game_mlink.find("a")['href']+".html"
        sh2.cell(row=row,column=10).value=game_mlink
    
    #For new Row
    row=row+1
    wb2.save("Game Details.xlsx")

    
wb1=openpyxl.load_workbook("games_id_names.xlsx")
sh1=wb1.active
print(sh1.max_row)
wb2=openpyxl.Workbook()
sh2=wb2.active
sh2.cell(row=1,column=1).value="Game Title"
sh2.cell(row=1,column=2).value="Release Date"
sh2.cell(row=1,column=3).value="Publisher"
sh2.cell(row=1,column=4).value="Price"
sh2.cell(row=1,column=5).value="Languages"
sh2.cell(row=1,column=6).value="Descriptions"
sh2.cell(row=1,column=7).value="Mature Content Descriptions"
sh2.cell(row=1,column=8).value="System Requirements"
sh2.cell(row=1,column=9).value="Meta Critic Score"
sh2.cell(row=1,column=10).value="Meta Critic Score Link"
    
for i in range(2,(sh1.max_row+1)): 
    game_name=(sh1.cell(row=i,column=1)).value
    print(game_name)
    game_link=(sh1.cell(row=i,column=4)).value
    print(game_link)
    filename=game_name+" data.xlsx"
    print(filename)
    print("\n")
    retrive_data(filename,game_link)
print("Done")
