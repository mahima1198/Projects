import requests
from datetime import datetime
import pycountry
import json
import openpyxl
import time

############### VARIABLES ################

#steam API Call
review_url_samp="https://store.steampowered.com/appreviews/"
user_profile_url="http://api.steampowered.com/ISteamUser/GetPlayerSummaries/v0002/"
#param can be obtained from steam website
review_type="postive"
steamids=0
day_range=0
total_reviews=1000
ct=0
#function for retrieving reviews from steam  
def get_steam_reviews(review_url,filename,game_name):
    global ct,day_range,review_type
    ct=0
    day_range=0
    review_type="postive"
    review_url=review_url
    key='4F1B5E492903F0967512AD3AE03BA971'
    count_list=[]
    check_id=[]
    row=2
    wb2=openpyxl.Workbook()
    sh2=wb2.active
    sh2.cell(row=1,column=1).value="game"
    sh2.cell(row=1,column=2).value="user_id"
    sh2.cell(row=1,column=3).value="user_name"
    sh2.cell(row=1,column=4).value="user_country"
    sh2.cell(row=1,column=5).value="review"
    sh2.cell(row=1,column=6).value="review_type"
    sh2.cell(row=1,column=7).value="review_date"

    while(ct<=total_reviews):
        for x in range(0,2):
            if(review_type=="positive"):
                review_url_param={'json':1,'num_per_page':100,'day_range':day_range,'review_type':'positive'}
                r1=requests.get(url=review_url,params=review_url_param)
                review_data=r1.json()
                review_type="negative"
                user_review_type="positive"
            else:
                review_url_param={'json':1,'num_per_page':100,'day_range':day_range,'review_type':'negative'}
                r1=requests.get(url=review_url,params=review_url_param)
                review_data=r1.json()
                review_type="positive"
                user_review_type="negative"
                
            for i in range(0,len(review_data['reviews'])):
                user_steam_id=review_data['reviews'][i].get('author')['steamid']
                user_review=review_data['reviews'][i].get('review')
                user_review_timestamp=review_data['reviews'][i].get('timestamp_created')
                user_review_date=str(datetime.fromtimestamp(user_review_timestamp))
                #print(user_steam_id)
                if(user_steam_id not in check_id ):
                    check_id.append(user_steam_id)
                    sh2.cell(row=row,column=1).value=game_name
                    sh2.cell(row=row,column=2).value=user_steam_id
                    #sheet.cell(row=row,column=3).value=user_name
                    sh2.cell(row=row,column=5).value=user_review
                    sh2.cell(row=row,column=6).value=user_review_type
                    sh2.cell(row=row,column=7).value=(user_review_date.split(' ')[0])
                    wb2.save(filename)
                    ct=ct+1
                    row=row+1
                count_list.append(ct)
                #print(ct)
                #print("day range:",day_range)
                #print("Occur",count_list.count(ct))
                if(count_list.count(ct)==15):
                    count_list.clear()
                    print("count:",ct)
                    print("day_range",day_range)
                    day_range=int(day_range*1.2)
                    if(day_range>400000):
                        del wb2
                        del sh2
                        return None
        print("day_range",day_range)
        day_range=day_range+50
    return None

wb1=openpyxl.load_workbook("games_id_names.xlsx")
sh1=wb1.active
for i in range(15,16): 
    game_name=(sh1.cell(row=i,column=1)).value
    print(game_name)
    game_id=(sh1.cell(row=i,column=2)).value
    print(game_id)
    review_url=review_url_samp+str(game_id)
    filename=game_name+".xlsx"
    get_steam_reviews(review_url,filename,game_name)
    del review_url

