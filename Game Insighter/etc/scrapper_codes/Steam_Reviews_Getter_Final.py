import requests
from datetime import datetime
import pycountry
import json
import openpyxl
import time

############### VARIABLES ################

#steam API Call
review_url="https://store.steampowered.com/appreviews/730"
user_profile_url="http://api.steampowered.com/ISteamUser/GetPlayerSummaries/v0002/"
#param can be obtained from steam website
day_range=0
review_type="postive"
steamids=0
total_reviews=800
ct=0
workbook=openpyxl.load_workbook("/home/harshit/projects/gameinsighter/scrapper_codes/SteamReviews.xlsx")
sheet=workbook["Sheet"]
#function for retrieving reviews from steam  
def get_steam_reviews():
    key='4F1B5E492903F0967512AD3AE03BA971'
    count_list=[]
    check_id=[]
    row=2
    global ct,day_range,review_type
    while(ct<=total_reviews):
        for x in range(0,2):
            if(review_type=="positive"):
                review_url_param={'json':1,'num_per_page':100,'day_range':day_range,'review_type':'positive'}
                r1=requests.get(url=review_url,params=review_url_param)
                review_data=r1.json()
                review_type="negative"
                user_review_type="positive"
            else:
                review_url_param={'json':1,'num_per_page':100,'day_range':day_range,'review_type':'negative'}
                r1=requests.get(url=review_url,params=review_url_param)
                review_data=r1.json()
                review_type="positive"
                user_review_type="negative"
                
            for i in range(0,len(review_data['reviews'])):
                user_steam_id=review_data['reviews'][i].get('author')['steamid']
                user_review=review_data['reviews'][i].get('review')
                user_review_timestamp=review_data['reviews'][i].get('timestamp_created')
                user_review_date=str(datetime.fromtimestamp(user_review_timestamp))
                if(i==30 or i==60 or i==90):
                    time.sleep(5)
                #print(user_steam_id)
                if(user_steam_id not in check_id ):
                    sids=int(user_steam_id)
                    user_profile_param={'key':key,'steamids':sids}
                    r2=requests.get(url=user_profile_url,params=user_profile_param)
                    user_data=r2.json()
                    user_name=user_data['response']['players'][0].get('personaname')
                    user_country_code=user_data['response']['players'][0].get('loccountrycode')
                    check_id.append(user_steam_id)
                    #print(review_data)
                    if(user_country_code!=None):
                        x=pycountry.countries.get(alpha_2=user_country_code)
                        #print(x.name)
                        if(x==None):
                            sheet.cell(row=row,column=4).value="null"
                        else:
                            sheet.cell(row=row,column=4).value=x.name
                    else:
                        sheet.cell(row=row,column=4).value="null"
                    sheet.cell(row=row,column=1).value="Counter Strike Global Offensive"
                    sheet.cell(row=row,column=2).value=user_steam_id
                    sheet.cell(row=row,column=3).value=user_name
                    sheet.cell(row=row,column=5).value=user_review
                    sheet.cell(row=row,column=6).value=user_review_type
                    sheet.cell(row=row,column=7).value=(user_review_date.split(' ')[0])
                    workbook.save('SteamReviews.xlsx')
                    ct=ct+1
                    row=row+1
                count_list.append(ct)
                
                #print("day range:",day_range)
                #print("Occur",count_list.count(ct))
                if(count_list.count(ct)==15):
                    count_list.clear()
                    print("count:",ct)
                    print("day_range",day_range)
                    day_range=int(day_range*1.2)
        print("day_range",day_range)
        day_range=day_range+50
    return None


get_steam_reviews()

