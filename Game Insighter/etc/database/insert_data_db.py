import os
os.environ["DJANGO_SETTINGS_MODULE"]="gameinsighter.settings"
from gameinsighter.settings import *
import django
django.setup()

from webapp.models import gameDetails
from openpyxl import load_workbook

loc=load_workbook(filename="./webapp/static/webapp/Excel Files/Game Details.xlsx")
wb=loc.active
data=[]
for i in range(2,14): 
    game=gameDetails()
    game.Steam_ID = wb.cell(row=i,column=1).value
    game.Game_Title = wb.cell(row=i,column=2).value
    game.Genre =  wb.cell(row=i,column=3).value
    game.Release_Date = wb.cell(row=i,column=4).value
    game.Publisher = wb.cell(row=i,column=5).value
    game.Price = wb.cell(row=i,column=6).value
    game.Languages = wb.cell(row=i,column=7).value
    game.Descriptions = wb.cell(row=i,column=8).value
    game.Mature_Content_Descriptions = wb.cell(row=i,column=9).value
    game.System_Requirements = wb.cell(row=i,column=10).value
    game.Metric_Critic_Score = wb.cell(row=i,column=11).value
    game.Metric_Critic_Score_Link = wb.cell(row=i,column=12).value
    game.Links = wb.cell(row=i,column=13).value
    game.save()